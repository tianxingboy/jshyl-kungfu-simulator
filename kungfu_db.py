# -*- coding: utf-8 -*-
"""武功秘籍数据库解析 - 从武功秘籍.xlsx读取(52列完整结构)"""
import os
import sys

def resource_path(relative_path):
    """获取资源文件路径(兼容PyInstaller打包)"""
    if hasattr(sys, '_MEIPASS'):
        return os.path.join(sys._MEIPASS, relative_path)
    return os.path.join(os.path.dirname(os.path.abspath(__file__)), relative_path)

class Kungfu:
    def __init__(self):
        self.item_id = 0          # 物品序号
        self.name = ''            # 物品名称(秘籍名)
        self.skill_name = ''      # 练出武功
        self.desc = ''            # 物品说明
        self.category = ''        # 种类: 拳掌/剑法/兵器/指腿/暗毒/内功/轻功/其它
        self.tier = 0             # 秘籍品阶(按主要系数加成推断)

        # 修炼需求
        self.qualification = None  # 需资质(正数=≥, 负数=低于)
        self.need_skills = {}      # 需五系: {'拳掌':200, ...}
        self.limit_person = None     # 限使用人物1(编号, None=不限)
        self.limit_person2 = None    # 限使用人物2(编号, None=不限)
        self.gender = ''           # 需性别
        self.inner_type = ''       # 需内力性质(阳/阴/调和)
        self.need_inner = None     # 需内力
        self.need_attack = None    # 需攻击
        self.need_dodge = None     # 需轻功
        self.need_use_poison = None # 需用毒
        self.need_medical = None   # 需医疗
        self.need_family = None    # 需家族(编号)
        self.need_school = None    # 需门派(编号)

        # 每级加成(修炼×10, 洗武功×9)
        self.bonuses = {}          # {'攻击':3, '防御':2, '轻功':5, '拳掌':9, ...}
        self.change_inner = ''     # 改变内力性质
        self.attack_times = 0      # 攻击次数
        self.attack_poison = 0     # 加攻击带毒

        # 其他
        self.price = 0             # 出售价格
        self.material = ''         # 需材料
        self.newbie_bonus = False  # 新手光环计算

    def __repr__(self):
        return f'<Kungfu #{self.item_id} {self.name}->{self.skill_name} [{self.category}] +{self.tier}>'

    def get_total_bonus(self, level=10):
        """获取指定等级的总加成字典"""
        return {k: v * level for k, v in self.bonuses.items() if v != 0}

    def get_bonus_str(self):
        """获取加成描述字符串(每级)"""
        parts = []
        for k, v in self.bonuses.items():
            if v != 0:
                sign = '+' if v > 0 else ''
                parts.append(f'{k}{sign}{v}')
        return ' '.join(parts) if parts else '-'

CAT_MAP = {
    '拳掌': '拳掌', '剑法': '御剑', '御剑': '御剑', '兵器': '兵器',
    '指腿': '指腿', '暗毒': '暗毒', '内功': '内功',
    '轻功': '轻功', '其它': '其它',
}

# 加成列映射: (列索引0-based, 字段名)
BONUS_COLS = [
    (19, '生命'),    # 加生命
    (20, '血量'),    # 加血量
    (22, '内力'),    # 加内力
    (23, '功力'),    # 加功力
    (24, '攻击'),    # 加攻击
    (25, '轻功'),    # 加轻功
    (26, '防御'),    # 加防御
    (27, '医疗'),    # 加医疗
    (28, '用毒'),    # 加用毒
    (29, '解毒'),    # 加解毒
    (30, '抗毒'),    # 加抗毒
    (31, '拳掌'),    # 加拳掌
    (32, '御剑'),    # 加御剑
    (33, '兵器'),    # 加兵器
    (34, '指腿'),    # 加指腿
    (35, '暗毒'),    # 加暗毒
    (36, '武常'),    # 加武常
    (38, '攻击带毒'), # 加攻击带毒
]

# 需求列映射
NEED_COLS = [
    (6, '拳掌', 'need_skills'),
    (7, '御剑', 'need_skills'),
    (8, '兵器', 'need_skills'),
    (9, '指腿', 'need_skills'),
    (10, '暗毒', 'need_skills'),
]

def _safe_int(val, default=None):
    """安全转换为int"""
    if val is None:
        return default
    try:
        return int(val)
    except (ValueError, TypeError):
        return default

def load_kungfu_db(xlsx_path):
    """从武功秘籍.xlsx加载所有武功(52列结构)"""
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    kungfus = []

    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
        name = row[2]  # 物品名称
        if not name:
            continue

        kf = Kungfu()
        kf.item_id = _safe_int(row[1], 0)
        kf.name = str(name).strip()
        kf.skill_name = str(row[3]).strip() if row[3] else ''
        kf.desc = str(row[4]).strip() if row[4] else ''
        kf.category = CAT_MAP.get(str(row[0]).strip(), '其它') if row[0] else '其它'

        # 需求
        q = _safe_int(row[5])
        if q is not None:
            kf.qualification = q
        for col_idx, skill_name, target in NEED_COLS:
            v = _safe_int(row[col_idx])
            if v is not None:
                kf.need_skills[skill_name] = v
        if row[11]:  # 限使用人物
            kf.limit_person = str(row[11]).strip()
        if row[12]:  # 需性别
            kf.gender = str(row[12]).strip()
        if row[13]:  # 需内力性质
            kf.inner_type = str(row[13]).strip()
        kf.need_inner = _safe_int(row[14])
        kf.need_attack = _safe_int(row[15])
        kf.need_dodge = _safe_int(row[16])
        kf.need_use_poison = _safe_int(row[17])
        kf.need_medical = _safe_int(row[18])

        # 加成(每级)
        for col_idx, field_name in BONUS_COLS:
            v = _safe_int(row[col_idx])
            if v is not None and v != 0:
                kf.bonuses[field_name] = v

        if row[21]:  # 改变内力性质
            kf.change_inner = str(row[21]).strip()
        at = _safe_int(row[37])
        if at is not None:
            kf.attack_times = at
        ap = _safe_int(row[38])
        if ap is not None:
            kf.attack_poison = ap
        kf.price = _safe_int(row[39], 0)
        if row[40]:  # 需材料
            kf.material = str(row[40]).strip()
        if row[51]:  # 新手光环计算
            kf.newbie_bonus = str(row[51]).strip() == '是'

        # 推断品阶: 五系系数加成的最大值
        skill_bonuses = [kf.bonuses.get(s, 0) for s in ['拳掌', '御剑', '兵器', '指腿', '暗毒']]
        kf.tier = max(skill_bonuses) if skill_bonuses else 0

        kungfus.append(kf)

    # 按品阶从小到大排序
    kungfus.sort(key=lambda k: (k.tier, k.category, k.item_id))
    return kungfus

def get_categories():
    return ['拳掌', '御剑', '兵器', '指腿', '暗毒', '内功', '轻功', '特殊']

if __name__ == '__main__':
    db = load_kungfu_db(r'E:\game\金书红颜录修改\jshyl5.60版\武功秘籍.xlsx')
    print(f'共{len(db)}条秘籍')
    # 按品阶统计
    from collections import Counter
    tiers = Counter(k.tier for k in db)
    print(f'品阶分布: {dict(sorted(tiers.items()))}')
    # 显示几条
    for kf in db[:5]:
        print(f'  [{kf.category}] +{kf.tier} {kf.name}->{kf.skill_name} 资{kf.qualification} 加成:{kf.bonuses}')
    # 找雪遁和太极拳
    for kf in db:
        if '雪遁' in kf.name or '太极拳经' in kf.name:
            print(f'  >>> [{kf.category}] +{kf.tier} {kf.name}->{kf.skill_name} 资{kf.qualification} 加成:{kf.bonuses} 需求:{kf.need_skills}')
